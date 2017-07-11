import requests
import mysql_database
import time
import random
from openpyxl import Workbook
import json
def insert_profile(term, gre_date, work_experience_months, ug_score, ug_score_pattern, en_exam_pattern,
                   en_exam_score,
                   gre_score, status, course_name, university_name, gre_verbal_score, gre_quant_score,
                   gre_awa_score, en_exam_writing_score, en_exam_listening_score,
                   en_exam_reading_score, en_exam_speaking_score, nickname):
    term = str(term)
    gre_date = str(gre_date)
    work_experience_months = float(work_experience_months)
    ug_score = float(ug_score)
    ug_score_pattern = str(ug_score_pattern)
    en_exam_pattern = str(en_exam_pattern)
    en_exam_score = float(en_exam_score)
    gre_score = int(gre_score)
    status = str(status)
    course_name = str(course_name)
    university_name = str(university_name)
    gre_verbal_score = int(gre_verbal_score)
    gre_quant_score = int(gre_quant_score)
    gre_awa_score = float(gre_awa_score)
    en_exam_writing_score = int(en_exam_writing_score)
    en_exam_listening_score = int(en_exam_listening_score)
    en_exam_reading_score = int(en_exam_reading_score)
    en_exam_speaking_score = int(en_exam_speaking_score)
    nickname = str(nickname)
    emp_data = dict(term=term, gre_date=gre_date, work_experience_months=work_experience_months, ug_score=ug_score,
                    ug_score_pattern=ug_score_pattern, en_exam_pattern=en_exam_pattern, en_exam_score=en_exam_score,
                    gre_score=gre_score, status=status, course_name=course_name, university_name=university_name,
                    gre_verbal_score=gre_verbal_score, gre_quant_score=gre_quant_score, gre_awa_score=gre_awa_score,
                    en_exam_writing_score=en_exam_writing_score, en_exam_listening_score=en_exam_listening_score,
                    en_exam_reading_score=en_exam_reading_score, en_exam_speaking_score=en_exam_speaking_score,
                    nickname=nickname)
    print emp_data
    return emp_data

def start():

    with open('H:\\Python_Programs\\yocket\\initial_data.json') as data_file:
        data = json.load(data_file)
    header = {
        'Accept-Encoding':'gzip',
        'Connection': 'Keep-Alive',
        'Content-Type':'application/x-www-form-urlencoded',
        'Content-Length':'255',
        'Host':'yocket.in',
        'Authorization' : 'Basic bXVrdW5kd2FnaEBnbWFpbC5jb206NDc2N2QzNjBjODE3Y2NkOWVmZGRjOWFlN2JlYzI5ODVmZDFiZDJiMw==',
        'User-Agent':'Dalvik/2.1.0 (Linux; U; Android 6.0.1; Redmi Note 3 MIUI/V8.2.3.0.MHOMIDL'
    }
    """payload={
    "user_id":301629,
    "mobile_device_id":861375035532500,
    "is_online":1
    }
    """
    payload={"user_id":"",
    "applications":data.get('applications'),
    "gre_score":"",
    "term":"",
    "year":"",
    "gre_date[year]":"",
    "gre_date[month]":"",
    "gre_date[day]":"",
    "en_exam_key":"",
    "en_exam_score":"",
    "ug_score":"",
    "ug_score_key":"",
    "work_experience_months":"",
    "tpp_key":"",
    "visa_status":"",
    "interested_course_group_id":"",
    "application_status":2
    }
    workbook = Workbook()
    worksheet = workbook.active

    try:

        r = requests.post(
            'https://yocket.in:443/user_profiles/find/matching_applicants.json?page=1',headers=header,
            data=payload)
        print r.json()
        row = 1

        for pagecount in range(1,r.json().get('pageCount')):
            data_url = 'https://yocket.in:443/user_profiles/find/matching_applicants.json?page=%s' %(pagecount)
            time.sleep(random.randint(0, 5))
            r = requests.post(
                data_url,
                headers=header,
                data=payload)
            for u in r.json().get('userProfiles'):
                try:
                    user_url="https://yocket.in:443/users/view/"+str(u.get('user').get('id'))+".json"
                    user_header = {
                        'Accept-Encoding': 'gzip',
                        'Connection': 'Keep-Alive',
                        'Host': 'yocket.in',
                        'Authorization': 'Basic bXVrdW5kd2FnaEBnbWFpbC5jb206NDc2N2QzNjBjODE3Y2NkOWVmZGRjOWFlN2JlYzI5ODVmZDFiZDJiMw==',
                        'User-Agent': 'Dalvik/2.1.0 (Linux; U; Android 6.0.1; Redmi Note 3 MIUI/V8.2.3.0.MHOMIDL'
                    }
                    time.sleep(random.randint(0,4))
                    user = requests.get(user_url,headers=user_header)
                    emp_data=insert_profile("none" if u.get('profile').get('term',"none")is None else u.get('profile').get('term',"none"),
                                                  "none" if u.get('profile').get('gre_date',"none")is None else u.get('profile').get('gre_date',"none"),
                                                  0 if u.get('profile').get('work_experience_months') is None else u.get('profile').get('work_experience_months'),
                                                  0 if u.get('profile').get('ug_score')is None else u.get('profile').get('ug_score'),
                                                  "none" if u.get('profile').get('ug_score_pattern',"none")is None else u.get('profile').get('ug_score_pattern',"none"),
                                                  "none" if u.get('profile').get('en_exam_pattern',"none")is None else u.get('profile').get('en_exam_pattern',"none"),
                                                  0 if u.get('profile').get('en_exam_score')==(None or 'NA') else u.get('profile').get('en_exam_score'),
                                                  0 if u.get('profile').get('gre_score')is None else u.get('profile').get('gre_score'),
                                                  "none" if u.get('application').get('status',"none")is None else u.get('application').get('status',"none"),
                                                  "none" if u.get('application').get('course_name',"none")is None else u.get('application').get('course_name',"none"),
                                                  "none" if u.get('application').get('university_name',"none")is None else u.get('application').get('university_name',"none"),
                                                  0 if user.json().get('user').get('user_profile').get('gre_verbal_score')is None else user.json().get('user').get('user_profile').get('gre_verbal_score'),
                                                  0 if user.json().get('user').get('user_profile').get('gre_quant_score')is None else user.json().get('user').get('user_profile').get('gre_quant_score'),
                                                  0 if user.json().get('user').get('user_profile').get('gre_awa_score')is None else user.json().get('user').get('user_profile').get('gre_awa_score'),
                                                  0 if user.json().get('user').get('user_profile').get('en_exam_writing_score')is None else user.json().get('user').get('user_profile').get('en_exam_writing_score'),
                                                  0 if user.json().get('user').get('user_profile').get('en_exam_listening_score')is None else user.json().get('user').get('user_profile').get('en_exam_listening_score'),
                                                  0 if user.json().get('user').get('user_profile').get('en_exam_reading_score')is None else user.json().get('user').get('user_profile').get('en_exam_reading_score'),
                                                  0 if user.json().get('user').get('user_profile').get('en_exam_speaking_score')is None else user.json().get('user').get('user_profile').get('en_exam_speaking_score'),
                                                  "none" if user.json().get('user').get('nickname') is None else user.json().get('user').get('nickname')
                                                  )
                    co = 1
                    for k, v in emp_data.items():
                        if row ==1:
                            worksheet.cell(row=row, column=co).value = k
                            co=co+1
                    row = row +1
                    col = 1
                    for k, v in emp_data.items():
                        #print row, col,v
                        worksheet.cell(row=row,column=col).value= v
                        col = col + 1
                    row = row + 1
                    workbook.save(data.get('file_path'))
                except Exception as e:
                    print "An error has occured" + e.message
                    import traceback
                    traceback.print_exc()
            print "data inserted at page %d" %(pagecount)
    finally:
        pass

if __name__=="__main__":
    start()